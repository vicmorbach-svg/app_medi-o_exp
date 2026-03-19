import subprocess
import platform
from pathlib import Path
from config import OUTPUT_DIR


def excel_para_pdf(excel_path: Path, aba: str) -> Path:
    """
    Converte uma aba do Excel para PDF.
    - Windows: usa xlwings + Excel instalado (COM)
    - Linux/Mac: usa LibreOffice headless
    """
    nome_pdf  = f"{excel_path.stem}_{aba}.pdf"
    pdf_path  = OUTPUT_DIR / nome_pdf
    sistema   = platform.system()

    if sistema == "Windows":
        _converter_windows(excel_path, aba, pdf_path)
    else:
        _converter_libreoffice(excel_path, aba, pdf_path)

    return pdf_path


def _converter_windows(excel_path: Path, aba: str, pdf_path: Path):
    import xlwings as xw
    with xw.App(visible=False) as app:
        wb = app.books.open(str(excel_path.resolve()))
        sheet = wb.sheets[aba]
        sheet.activate()
        wb.api.ActiveSheet.ExportAsFixedFormat(
            Type=0,
            Filename=str(pdf_path.resolve()),
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
        )
        wb.close()


def _converter_libreoffice(excel_path: Path, aba: str, pdf_path: Path):
    """
    LibreOffice não exporta abas individualmente via linha de comando simples.
    Estratégia: cria uma cópia temporária só com a aba desejada, converte, apaga.
    """
    import shutil
    import openpyxl

    tmp_path = OUTPUT_DIR / f"_tmp_{aba}_{excel_path.name}"

    # Cria workbook temporário só com a aba desejada
    wb_orig = openpyxl.load_workbook(excel_path)
    wb_tmp  = openpyxl.Workbook()
    wb_tmp.remove(wb_tmp.active)

    ws_orig = wb_orig[aba]
    from copy import copy

    ws_new = wb_tmp.create_sheet(aba)
    for row in ws_orig.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font      = copy(cell.font)
                new_cell.border    = copy(cell.border)
                new_cell.fill      = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)

    wb_tmp.save(tmp_path)

    cmd = [
        "libreoffice", "--headless",
        "--convert-to", "pdf",
        "--outdir", str(OUTPUT_DIR),
        str(tmp_path.resolve())
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    tmp_path.unlink(missing_ok=True)

    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice erro: {result.stderr}")

    gerado = OUTPUT_DIR / (tmp_path.stem + ".pdf")
    if gerado.exists():
        gerado.rename(pdf_path)


def gerar_pdfs_medicao(excel_path: Path) -> dict:
    """Gera PDF para PROTOCOLO e BOLETIM. Retorna {'PROTOCOLO': Path, 'BOLETIM': Path}"""
    return {
        "PROTOCOLO": excel_para_pdf(excel_path, "PROTOCOLO"),
        "BOLETIM":   excel_para_pdf(excel_path, "BOLETIM"),
    }
