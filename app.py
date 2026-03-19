import streamlit as st
import pandas as pd
from datetime import date
from pathlib import Path
import hashlib # Para hash de senhas (segurança básica)
import shutil # Para copiar o arquivo modelo
import openpyxl # Para manipular o Excel
import io # Para manipular arquivos em memória
import json # Para lidar com a string JSON de preços

# ── Configurações de diretórios ───────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DADOS_DIR = BASE_DIR / "dados"
MODELO_DIR = BASE_DIR / "modelo"
OUTPUT_DIR = BASE_DIR / "output"

for d in [DADOS_DIR, OUTPUT_DIR]:
    d.mkdir(exist_ok=True)

CONTRATOS_FILE = DADOS_DIR / "contratos.xlsx"
MEDICOES_FILE  = DADOS_DIR / "medicoes.xlsx"
MODELO_FILE    = MODELO_DIR / "Modelo_medio.xlsx" # Seu modelo com as abas renomeadas e aba DADOS

# ── Funções de utilidade ─────────────────────────────────────────────────────
def fmt_brl(value):
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def to_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, date):
        return value
    try:
        return pd.to_datetime(value).date()
    except:
        return None

def extrair_mes_do_periodo(periodo: str) -> str:
    MESES_PT = {
        1: "JANEIRO",  2: "FEVEREIRO", 3: "MARÇO",    4: "ABRIL",
        5: "MAIO",     6: "JUNHO",     7: "JULHO",     8: "AGOSTO",
        9: "SETEMBRO", 10: "OUTUBRO",  11: "NOVEMBRO", 12: "DEZEMBRO"
    }
    try:
        partes = periodo.upper().strip().split("A")
        data_fim = partes[-1].strip()
        segmentos = data_fim.replace("-", "/").split("/")
        if len(segmentos) >= 3:
            mes = int(segments[1])
            ano = segments[2][:4]
            return f"{MESES_PT.get(mes, '')} {ano}"
        elif len(segments) == 2:
            mes = int(segments[1])
            return MESES_PT.get(mes, "")
    except Exception:
        pass
    return ""

# ── Funções de carregamento/salvamento de dados ──────────────────────────────
@st.cache_data
def carregar_contratos():
    if CONTRATOS_FILE.exists():
        df = pd.read_excel(CONTRATOS_FILE)
        if 'fornecedor' not in df.columns:
            df['fornecedor'] = ''
        # Converte a string JSON de precos_servicos para um dicionário
        if 'precos_servicos' in df.columns:
            df['precos_servicos'] = df['precos_servicos'].apply(lambda x: json.loads(x) if pd.notna(x) else {})
        else:
            df['precos_servicos'] = {} # Garante que a coluna exista
        return df
    return pd.DataFrame(columns=[
        "contrato", "empresa", "local", "modalidade", "data_base",
        "data_termino", "valor_original", "item_num", "und",
        "quant_total", "servicos_disponiveis", "precos_servicos", # Atualizado
        "centro_custo", "conta_contabil", "item_caixa", "fornecedor"
    ])

@st.cache_data
def carregar_medicoes():
    if MEDICOES_FILE.exists():
        return pd.read_excel(MEDICOES_FILE)
    return pd.DataFrame(columns=[
        "contrato", "num_medicao", "empresa", "local", "modalidade",
        "data_base", "data_termino", "periodo", "mes_execucao",
        "data_apresentacao", "descricao_servico", "item_num", "und",
        "quant_total", "preco_unitario", # Manter para compatibilidade com histórico
        "quant_mes", "valor_mes",
        "quant_acum_ant", "valor_acum_ant", "quant_acum_total",
        "valor_acum_total", "valor_original", "saldo_contrato",
        "centro_custo", "conta_contabil", "item_caixa"
    ])

def salvar_medicoes(df: pd.DataFrame):
    df.to_excel(MEDICOES_FILE, index=False)
    carregar_medicoes.clear() # Limpa cache para recarregar

def calcular_acumulado(df_medicoes_filtrado, contrato_sel, num_medicao_atual):
    medicoes_anteriores = df_medicoes_filtrado[
        (df_medicoes_filtrado["contrato"].astype(str) == str(contrato_sel)) &
        (df_medicoes_filtrado["num_medicao"].astype(int) < int(num_medicao_atual))
    ]
    quant_acum_ant = medicoes_anteriores["quant_mes"].sum()
    valor_acum_ant = medicoes_anteriores["valor_mes"].sum()
    return quant_acum_ant, valor_acum_ant

# ── Funções de geração de Excel (nova abordagem) ─────────────────────────────
def gerar_excel_com_dados(df_medicoes_fornecedor: pd.DataFrame, contrato_selecionado: str, num_medicao_selecionada: int) -> io.BytesIO:
    """
    Copia o Modelo_medio.xlsx, preenche a aba DADOS com as medições do fornecedor
    e define as células de controle na aba PROTOCOLO para a medição selecionada.
    Retorna o arquivo Excel em memória (BytesIO).
    """
    if not MODELO_FILE.exists():
        st.error(f"Arquivo modelo não encontrado em: {MODELO_FILE}")
        st.stop()

    # Copia o modelo para um buffer em memória
    template_buffer = io.BytesIO(MODELO_FILE.read_bytes())
    wb = openpyxl.load_workbook(template_buffer)

    # Preenche a aba DADOS
    if "DADOS" not in wb.sheetnames:
        ws_dados = wb.create_sheet("DADOS")
    else:
        ws_dados = wb["DADOS"]
        # Limpa conteúdo existente na aba DADOS (exceto cabeçalho)
        if ws_dados.max_row > 1:
            ws_dados.delete_rows(2, ws_dados.max_row - 1)

    # Escreve o cabeçalho (se a aba DADOS estava vazia ou foi limpa)
    if ws_dados.max_row < 1 or ws_dados['A1'].value is None:
        for col_idx, col_name in enumerate(df_medicoes_fornecedor.columns, 1):
            ws_dados.cell(row=1, column=col_idx, value=col_name)

    # Escreve os dados
    # Começa da linha 2 (abaixo do cabeçalho)
    for r_idx, row_data in df_medicoes_fornecedor.iterrows():
        for c_idx, value in enumerate(row_data, 1):
            ws_dados.cell(row=r_idx + 2, column=c_idx, value=value) # +2 porque linha 1 é cabeçalho, linha 2 é a primeira linha de dados

    # Define as células de controle na aba PROTOCOLO para a medição selecionada
    # Assumindo que B9 é o contrato e B5 é o num_medicao
    # E que as abas foram renomeadas para "PROTOCOLO" e "BOLETIM"
    if "PROTOCOLO" in wb.sheetnames:
        ws_protocolo = wb["PROTOCOLO"]
        ws_protocolo["B9"] = contrato_selecionado # Define o contrato para as fórmulas buscarem
        ws_protocolo["B5"] = num_medicao_selecionada # Define o num_medicao para as fórmulas buscarem
    else:
        st.warning("Aba 'PROTOCOLO' não encontrada no modelo. Verifique o nome da aba.")

    if "BOLETIM" in wb.sheetnames:
        # A aba BOLETIM geralmente referencia PROTOCOLO, então não precisa de ajuste direto aqui
        pass
    else:
        st.warning("Aba 'BOLETIM' não encontrada no modelo. Verifique o nome da aba.")

    # Salva o workbook em um buffer em memória
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0) # Volta para o início do buffer
    return output_buffer

# ── Autenticação (usando st.secrets) ─────────────────────────────────────────
def autenticar(username, password):
    if "usuarios" not in st.secrets:
        st.error("Configuração de usuários não encontrada em `secrets.toml`.")
        return False

    usuarios_secrets = st.secrets["usuarios"]
    hashed_password = hashlib.sha256(password.encode()).hexdigest()

    for user_data in usuarios_secrets:
        if user_data["usuario"] == username and hashlib.sha256(user_data["senha"].encode()).hexdigest() == hashed_password:
            st.session_state.logged_in = True
            st.session_state.username = username
            st.session_state.fornecedor = user_data["fornecedor"]
            st.session_state.is_admin = user_data.get("is_admin", False) # Pega o status de admin
            return True
    return False

def logout():
    st.session_state.logged_in = False
    st.session_state.username = None
    st.session_state.fornecedor = None
    st.session_state.is_admin = False
    st.rerun()

# ── Inicialização do estado da sessão ────────────────────────────────────────
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "excel_buffer" not in st.session_state: # Armazena o buffer do Excel gerado
    st.session_state.excel_buffer = None
if "excel_filename" not in st.session_state: # Armazena o nome do arquivo Excel gerado
    st.session_state.excel_filename = None
if "dados_gerados" not in st.session_state:
    st.session_state.dados_gerados = None
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

# ── Interface do Streamlit ───────────────────────────────────────────────────
st.set_page_config(layout="wide", page_title="App de Medição CORSAN")

if not st.session_state.logged_in:
    st.title("Login - App de Medição CORSAN")
    username = st.text_input("Usuário")
    password = st.text_input("Senha", type="password")
    if st.button("Entrar"):
        if autenticar(username, password):
            st.success(f"Bem-vindo, {st.session_state.username} ({st.session_state.fornecedor})!")
            st.rerun()
        else:
            st.error("Usuário ou senha incorretos.")
else:
    st.sidebar.title(f"Bem-vindo, {st.session_state.username}")
    st.sidebar.write(f"Fornecedor: **{st.session_state.fornecedor}**")
    if st.session_state.is_admin:
        st.sidebar.info("Você é um usuário administrador.")
    if st.sidebar.button("Sair"):
        logout()

    st.title("App de Medição CORSAN")

    # Carrega dados
    df_contratos = carregar_contratos()
    df_medicoes = carregar_medicoes()

    # Filtra dados pelo fornecedor logado
    df_contratos_fornecedor = df_contratos[
        df_contratos["fornecedor"] == st.session_state.fornecedor
    ].copy()
    df_medicoes_fornecedor = df_medicoes[
        df_medicoes["empresa"] == st.session_state.fornecedor # Assumindo que 'empresa' em medicoes.xlsx é o fornecedor
    ].copy()

    if df_contratos_fornecedor.empty:
        st.warning(f"Nenhum contrato encontrado para o fornecedor '{st.session_state.fornecedor}'. Verifique o arquivo 'contratos.xlsx'.")
        st.stop()

    # ──────────────────────────────────────────────────────────────────────────────
    # 1. CONTRATO
    # ──────────────────────────────────────────────────────────────────────────────
    st.subheader("1. Seleção do Contrato")

    contratos_disp = df_contratos_fornecedor["contrato"].unique()
    contrato_sel = st.selectbox(
        "Selecione o Contrato",
        options=contratos_disp,
        format_func=lambda x: f"{x} - {df_contratos_fornecedor[df_contratos_fornecedor['contrato'] == x]['empresa'].iloc[0]}"
    )

    if not contrato_sel:
        st.info("Selecione um contrato para continuar.")
        st.stop()

    cr = df_contratos_fornecedor[df_contratos_fornecedor["contrato"] == contrato_sel].iloc[0]

    st.json(cr.to_dict()) # Exibe os detalhes do contrato selecionado

    st.divider()

    # ──────────────────────────────────────────────────────────────────────────────
    # 2. TIPO DE SERVIÇO E PREÇO UNITÁRIO
    # ──────────────────────────────────────────────────────────────────────────────
    st.subheader("2. Tipo de Serviço e Preço Unitário")

    servicos_precos = cr.get("precos_servicos", {})
    servicos_lista = list(servicos_precos.keys())

    if not servicos_lista:
        st.error("Nenhum serviço configurado para este contrato. Verifique 'contratos.xlsx'.")
        st.stop()

    descricao_servico = st.selectbox("Selecione o Tipo de Serviço", servicos_lista)
    preco_unitario_selecionado = servicos_precos.get(descricao_servico, 0.0)

    st.info(f"Preço Unitário para '{descricao_servico}': **{fmt_brl(preco_unitario_selecionado)}**")

    st.divider()

    # ──────────────────────────────────────────────────────────────────────────────
    # 3. DADOS DA MEDIÇÃO
    # ──────────────────────────────────────────────────────────────────────────────
    st.subheader("3. Dados da Medição")

    col1, col2, col3 = st.columns([1, 2, 2])

    with col1:
        # Sugere o próximo número de medição
        ult_medicao = df_medicoes_fornecedor[
            df_medicoes_fornecedor["contrato"].astype(str) == str(contrato_sel)
        ]["num_medicao"].max()
        prox_num_medicao = (ult_medicao + 1) if pd.notna(ult_medicao) else 1
        num_medicao = st.number_input("Nº do Boletim", min_value=1, step=1, value=int(prox_num_medicao))

    with col2:
        periodo = st.text_input(
            "Período",
            placeholder="01/02/2026 A 28/02/2026",
            help="Formato: DD/MM/AAAA A DD/MM/AAAA"
        )

    with col3:
        data_apresentacao = st.date_input("Data de apresentação", value=date.today())

    mes_execucao = extrair_mes_do_periodo(periodo) if periodo else ""

    if mes_execucao:
        st.success(f"Mês de execução: **{mes_execucao}**")
    elif periodo:
        st.warning("Não foi possível extrair o mês. Verifique o formato do período.")

    st.divider()

    # ──────────────────────────────────────────────────────────────────────────────
    # 4. QUANTIDADE
    # ──────────────────────────────────────────────────────────────────────────────
    st.subheader("4. Quantidade Medida")

    quant_mes = st.number_input(
        f"Quantidade medida no mês ({cr['und']})",
        min_value=0.0, step=0.01, value=0.0, format="%.2f"
    )

    # ── Cálculos ─────────────────────────────────────────────────────────────────
    valor_original   = float(cr["valor_original"])
    # preco_unitario agora vem da seleção do serviço
    preco_unitario   = preco_unitario_selecionado

    quant_acum_ant, valor_acum_ant = calcular_acumulado(
        df_medicoes_fornecedor, contrato_sel, num_medicao
    )

    valor_mes        = quant_mes * preco_unitario
    quant_acum_total = quant_acum_ant + quant_mes
    valor_acum_total = valor_acum_ant + valor_mes
    saldo_contrato   = valor_original - valor_acum_total

    st.divider()

    # ──────────────────────────────────────────────────────────────────────────────
    # 5. RESUMO
    # ──────────────────────────────────────────────────────────────────────────────
    st.subheader("5. Resumo da Medição")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Acum. Anterior",   fmt_brl(valor_acum_ant))
    c2.metric("Valor do Mês",     fmt_brl(valor_mes))
    c3.metric("Acum. Total",      fmt_brl(valor_acum_total))
    c4.metric("Saldo do Contrato", fmt_brl(saldo_contrato))

    pct = (valor_acum_total / valor_original * 100) if valor_original > 0 else 0
    st.progress(
        min(pct / 100, 1.0),
        text=f"Execução contratual: {pct:.1f}%"
    )

    st.divider()

    # ──────────────────────────────────────────────────────────────────────────────
    # 6. GERAR E SALVAR ARQUIVOS
    # ──────────────────────────────────────────────────────────────────────────────
    st.subheader("6. Gerar e Salvar Arquivos")

    campos_ok = bool(periodo and mes_execucao and quant_mes >= 0 and descricao_servico)

    if not campos_ok:
        st.info("Preencha todos os campos obrigatórios para habilitar a geração.")

    if st.button("💾 Salvar Medição e Gerar Excel", type="primary",
                 use_container_width=True, disabled=not campos_ok):

        dados_medicao = {
            "contrato":           contrato_sel,
            "num_medicao":        num_medicao,
            "periodo":            periodo,
            "data_apresentacao":  data_apresentacao,
            "descricao_servico":  descricao_servico,
            "valor_mes":          valor_mes,
            "quant_mes":          quant_mes,
            "empresa":            cr["empresa"], # Empresa do contrato, que é o fornecedor logado
            "local":              cr["local"],
            "modalidade":         cr["modalidade"],
            "data_base":          to_date(cr["data_base"]),
            "data_termino":       to_date(cr["data_termino"]),
            "valor_original":     valor_original,
            "item_num":           cr["item_num"],
            "und":                cr["und"],
            "quant_total":        float(cr["quant_total"]),
            "preco_unitario":     preco_unitario, # Agora é o preço do serviço selecionado
            "centro_custo":       str(cr.get("centro_custo", "")),
            "conta_contabil":     str(cr.get("conta_contabil", "")),
            "item_caixa":         str(cr.get("item_caixa", "")),
            "quant_acum_ant":     quant_acum_ant,
            "valor_acum_ant":     valor_acum_ant,
            "quant_acum_total":   quant_acum_total,
            "valor_acum_total":   valor_acum_total,
            "saldo_contrato":     saldo_contrato,
        }

        # 1. Salvar a medição na base de dados geral (medicoes.xlsx)
        with st.spinner("Salvando medição..."):
            try:
                # Remove duplicata se existir
                mask = ~(
                    (df_medicoes["contrato"].astype(str) == str(contrato_sel)) &
                    (df_medicoes["num_medicao"].astype(int) == int(num_medicao))
                )
                df_medicoes_atualizado = df_medicoes[mask]

                novo_registro_df = pd.DataFrame([dados_medicao])
                df_medicoes_atualizado = pd.concat(
                    [df_medicoes_atualizado, novo_registro_df],
                    ignore_index=True
                )
                salvar_medicoes(df_medicoes_atualizado)
                st.success("Medição salva com sucesso na base de dados!")
                st.session_state.dados_gerados = dados_medicao # Armazena para download
            except Exception as e:
                st.error(f"Erro ao salvar medição: {e}")
                st.stop()

        # 2. Gerar o Excel final com os dados do fornecedor e a medição selecionada
        with st.spinner("Gerando Excel final com dados do fornecedor..."):
            try:
                # Recarrega as medições para incluir a recém-salva
                df_medicoes_atualizado_fornecedor = carregar_medicoes()[
                    carregar_medicoes()["empresa"] == st.session_state.fornecedor
                ].copy()

                excel_buffer = gerar_excel_com_dados(
                    df_medicoes_atualizado_fornecedor,
                    contrato_sel,
                    num_medicao
                )

                # Nome do arquivo PDF: 'fornecedor' + 'nome da aba' + 'serviço' + 'mês'
                # Assumindo que a aba principal é "PROTOCOLO"
                nome_excel_final = (
                    f"{st.session_state.fornecedor.replace(' ', '_')}_"
                    f"PROTOCOLO_{descricao_servico.replace(' ', '_')}_"
                    f"{mes_execucao.replace(' ', '_')}.xlsx"
                )
                st.session_state.excel_buffer = excel_buffer
                st.session_state.excel_filename = nome_excel_final
                st.success(f"Excel final gerado: `{nome_excel_final}`")
            except Exception as e:
                st.error(f"Erro ao gerar Excel final: {e}")
                st.stop()

# ──────────────────────────────────────────────────────────────────────────────
# 7. DOWNLOADS
# ──────────────────────────────────────────────────────────────────────────────
    if st.session_state.excel_buffer:
        st.divider()
        st.subheader("7. Downloads")

        col1, col2 = st.columns(2)

        col1.download_button(
            "📥 Baixar Excel da Medição (Modelo Preenchido)",
            data=st.session_state.excel_buffer,
            file_name=st.session_state.excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        col1.info("Para gerar o PDF, abra o arquivo Excel baixado e use a função 'Salvar como PDF' do próprio Excel.")

        # Opcional: permitir baixar a base de medicoes.xlsx completa do fornecedor
        df_medicoes_fornecedor_download = carregar_medicoes()[
            carregar_medicoes()["empresa"] == st.session_state.fornecedor
        ].copy()

        excel_buffer_medicoes = io.BytesIO()
        df_medicoes_fornecedor_download.to_excel(excel_buffer_medicoes, index=False, engine="openpyxl")
        excel_buffer_medicoes.seek(0)

        col2.download_button(
            "📥 Baixar Base de Medições do Fornecedor (medicoes.xlsx)",
            data=excel_buffer_medicoes,
            file_name=f"medicoes_{st.session_state.fornecedor.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    # ──────────────────────────────────────────────────────────────────────────────
    # 8. EXCLUSÃO DE LANÇAMENTOS (APENAS ADMIN)
    # ──────────────────────────────────────────────────────────────────────────────
    if st.session_state.is_admin:
        st.divider()
        st.subheader("8. Excluir Lançamento (Apenas Admin)")
        st.warning("Esta funcionalidade é apenas para administradores e remove dados permanentemente.")

        # Exibe as medições do fornecedor logado para seleção
        if not df_medicoes_fornecedor.empty:
            st.dataframe(df_medicoes_fornecedor.sort_values(by=["contrato", "num_medicao"]), use_container_width=True)

            medicao_para_excluir = st.selectbox(
                "Selecione a Medição para Excluir",
                options=df_medicoes_fornecedor.apply(
                    lambda row: f"Contrato: {row['contrato']} - Boletim: {row['num_medicao']} - Serviço: {row['descricao_servico']} - Mês: {row['mes_execucao']}",
                    axis=1
                ).tolist()
            )

            if medicao_para_excluir:
                # Extrai contrato e num_medicao da string selecionada
                partes = medicao_para_excluir.split(" - ")
                contrato_excluir = partes[0].replace("Contrato: ", "")
                num_medicao_excluir = int(partes[1].replace("Boletim: ", ""))
                descricao_servico_excluir = partes[2].replace("Serviço: ", "")

                if st.button(f"🔴 Confirmar Exclusão da Medição {num_medicao_excluir} do Contrato {contrato_excluir} ({descricao_servico_excluir})", type="secondary"):
                    try:
                        df_medicoes_atualizado = df_medicoes[
                            ~((df_medicoes["contrato"].astype(str) == contrato_excluir) &
                              (df_medicoes["num_medicao"].astype(int) == num_medicao_excluir) &
                              (df_medicoes["descricao_servico"].astype(str) == descricao_servico_excluir)) # Adiciona serviço para desambiguar
                        ].copy()
                        salvar_medicoes(df_medicoes_atualizado)
                        st.success(f"Medição {num_medicao_excluir} do Contrato {contrato_excluir} ({descricao_servico_excluir}) excluída com sucesso!")
                        st.rerun() # Recarrega a página para atualizar a lista
                    except Exception as e:
                        st.error(f"Erro ao excluir medição: {e}")
        else:
            st.info("Não há medições para excluir para este fornecedor.")
