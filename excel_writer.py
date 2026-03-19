import io
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from datetime import date
from config import MODELO_FILE, OUTPUT_DIR

MESES_PT = {
    1: "JANEIRO",  2: "FEVEREIRO", 3: "MARÇO",    4: "ABRIL",
    5: "MAIO",     6: "JUNHO",     7: "JULHO",     8: "AGOSTO",
    9: "SETEMBRO", 10: "OUTUBRO",  11: "NOVEMBRO", 12: "DEZEMBRO"
}

# ─────────────────────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill(
        start_color=hex_color, end_color=hex_color, fill_type="solid"
    )

def _thin_border(bottom_only=False):
    thin = Side(border_style="thin", color="000000")
    if bottom_only:
        return Border(bottom=thin)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _w(ws, cell, value):
    """Escreve valor na célula sem apagar formatação existente."""
    ws[cell] = value

def _wn(ws, row, col, value):
    """Escreve valor por índice numérico de linha/coluna."""
    ws.cell(row=row, column=col, value=value)

def _fmt_date(d):
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    try:
        import pandas as pd
        return pd.to_datetime(d).strftime("%d/%m/%Y")
    except Exception:
        return str(d)

# ─────────────────────────────────────────────────────────────
# EXTRAÇÃO DO MÊS DO PERÍODO
# ─────────────────────────────────────────────────────────────

def extrair_mes_do_periodo(periodo: str) -> str:
    """
    "01/02/2026 A 28/02/2026" → "FEVEREIRO 2026"
    "01/02 A 28/02/2026"      → "FEVEREIRO 2026"
    """
    try:
        partes = periodo.upper().strip().split("A")
        data_fim = partes[-1].strip()
        segmentos = data_fim.replace("-", "/").split("/")
        if len(segmentos) >= 3:
            mes = int(segmentos[1])
            ano = segmentos[2][:4]
            return f"{MESES_PT.get(mes, '')} {ano}"
        elif len(segmentos) == 2:
            mes = int(segmentos[1])
            return MESES_PT.get(mes, "")
    except Exception:
        pass
    return ""

# ─────────────────────────────────────────────────────────────
# CRIAÇÃO DO MODELO BASE (quando não existe arquivo físico)
# ─────────────────────────────────────────────────────────────

def _criar_aba_protocolo(wb: openpyxl.Workbook):
    ws = wb.create_sheet("PROTOCOLO")

    # Larguras de coluna
    for col, w in [("A",45),("B",40),("C",15),("D",30),
                   ("E",12),("F",12),("G",30),("H",20)]:
        ws.column_dimensions[col].width = w

    # Altura de linhas relevantes
    for row in range(1, 80):
        ws.row_dimensions[row].height = 18

    cinza   = _fill("D9D9D9")
    azul_cl = _fill("DDEEFF")

    def cabecalho(ws, row, texto, col="A", bold=True, fill=None):
        c = ws.cell(row=row, column=ord(col)-64, value=texto)
        c.font      = _font(bold=bold, size=10)
        c.alignment = _align()
        if fill:
            c.fill = fill
        return c

    # ── Linha 1: título ──────────────────────────────────────
    ws["B1"] = "PROTOCOLO DE MEDIÇÃO"
    ws["B1"].font      = _font(bold=True, size=12)
    ws["B1"].alignment = _align(h="center")
    ws["G1"] = "COMPANHIA RIOGRANDENSE DE SANEAMENTO - CORSAN"
    ws["G1"].font      = _font(bold=True, size=10)
    ws["G1"].alignment = _align(h="center")

    # ── Linha 3: contrato + endereço CORSAN ──────────────────
    ws["B3"] = "CONTRATO Nº"
    ws["B3"].font = _font(bold=True)
    # D3 = número do contrato (variável)
    ws["G3"] = "Rua Caldas Júnior nº. 120, 18º andar"

    # ── Linhas 4–5: complemento endereço ────────────────────
    ws["G4"] = "Porto Alegre/RS"
    ws["G5"] = "CNPJ: 92.802.784/0001-90"

    # ── Linha 6: número do boletim (variável A6) ─────────────
    ws["A6"].font      = _font(bold=True, size=11)
    ws["A6"].alignment = _align()

    # ── Linha 7: descrição do serviço (variável A7) ──────────
    ws["A7"].font      = _font(bold=False, size=10)
    ws["A7"].alignment = _align()

    # ── Linhas 8–12: dados do contrato ───────────────────────
    for label, row, col in [
        ("EMPRESA :",   8, "A"),
        ("CONTRATO :",  9, "A"),
        ("LOCAL:",     10, "A"),
        ("DATA BASE :", 11, "A"),
        ("PERÍODO :",  12, "A"),
    ]:
        c = ws.cell(row=row, column=1, value=label)
        c.font      = _font(bold=True)
        c.alignment = _align()

    ws["G9"]  = "MODALIDADE :"
    ws["G9"].font = _font(bold=True)
    ws["G12"] = "Término"
    ws["G12"].font = _font(bold=True)

    # ── Linhas 14–15: mês execução + data apresentação ───────
    ws["A14"] = "MÊS DE EXECUÇÃO DAS OBRAS/SERVIÇOS :"
    ws["A14"].font = _font(bold=True)
    ws["G14"] = "SOURCING"
    ws["G14"].font = _font(bold=True)
    ws["A15"] = "DATA DE APRESENTAÇÃO DA MEDIÇÃO :"
    ws["A15"].font = _font(bold=True)

    # ── Linha 17: 1.- VALOR DO CONTRATO JURÍDICO ─────────────
    ws["A17"] = "1.- VALOR DO CONTRATO JURÍDICO"
    ws["A17"].font = _font(bold=True, size=10)
    ws["A17"].fill = cinza

    # ── Linha 18: valor original ─────────────────────────────
    ws["B18"] = "VALOR ORIGINAL DO CONTRATO : R $"
    ws["B18"].font = _font(bold=True)

    # ── Linha 20: 2.- ADIANTAMENTO ───────────────────────────
    ws["A20"] = "2.- ADIANTAMENTO"
    ws["A20"].font = _font(bold=True)
    ws["A20"].fill = cinza
    ws["H20"] = 0

    ws["A21"] = "Data"
    ws["A21"].font = _font(bold=False, size=9, color="666666")

    # ── Linha 23: 3.- VALOR DA SOURCING ──────────────────────
    ws["A23"] = "3.- VALOR DA SOURCING"
    ws["A23"].font = _font(bold=True)
    ws["A23"].fill = cinza

    # ── Linha 28: 4.- MEDIÇÕES BRUTAS ────────────────────────
    ws["A28"] = "4.- MEDIÇÕES BRUTAS"
    ws["A28"].font = _font(bold=True)
    ws["A28"].fill = cinza

    # ── Linha 29: cabeçalho do histórico ─────────────────────
    ws["A29"] = "Data Aprovação"
    ws["B29"] = "Data Prev. Pagto"
    for col in ["A","B"]:
        ws[f"{col}29"].font      = _font(bold=True, size=9)
        ws[f"{col}29"].alignment = _align(h="center")
        ws[f"{col}29"].fill      = cinza

    # Linhas 30–44 ficam vazias (bloco amarelo — preenchido dinamicamente)

    # ── Seções fixas abaixo do histórico ─────────────────────
    # Linha 46: 5.- DESCONTO DE MATERIAIS
    ws["A46"] = "5.- DESCONTO DE MATERIAIS (FATURAMENTO DIRETO)"
    ws["A46"].font = _font(bold=True)
    ws["A46"].fill = cinza
    ws["H46"] = 0

    # Linha 51: 6.- RETENÇÕES
    ws["A51"] = "6.- RETENÇÕES CONTRATUAIS - 5%"
    ws["A51"].font = _font(bold=True)
    ws["A51"].fill = cinza
    for col_idx, val in [(5,0),(6,0),(7,0),(8,0)]:
        ws.cell(row=51, column=col_idx, value=val)

    # Linha 52: cabeçalho retenções
    for col, label in [("A","Data Retenção"),("B","Data Devolução"),
                        ("E","Valor Retido (R$)"),
                        ("G","Retenção (R) Devolução (D)"),
                        ("H","Resultado No Mês (R$)")]:
        ws[f"{col}52"] = label
        ws[f"{col}52"].font = _font(bold=True, size=9)
        ws[f"{col}52"].fill = cinza

    # Linha 54: 7.- OBSERVAÇÃO
    ws["A54"] = "7.-OBSERVAÇÃO"
    ws["A54"].font = _font(bold=True)
    ws["A54"].fill = cinza

    # ── Linhas de resultado (57, 59, 61, 63) ─────────────────
    # Serão calculadas dinamicamente no gerar_excel_medicao()
    # mas também deixamos os labels fixos para o modelo base:
    ws["A57"] = "VALOR BRUTO DA MEDIÇÃO NO MÊS (sem descontos)"
    ws["A57"].font = _font(bold=True)
    ws["A57"].fill = _fill("FFF2CC")  # amarelo claro

    ws["A59"] = "SALDO DO CONTRATO JURÍDICO"
    ws["A59"].font = _font(bold=True)
    ws["A59"].fill = _fill("E2EFDA")  # verde claro

    ws["A61"] = "SALDO DA SOURCING"
    ws["A61"].font = _font(bold=True)
    ws["A61"].fill = _fill("E2EFDA")

    ws["A63"] = "CENTRO DE CUSTO:"
    ws["C63"] = "CONTA CONTÁBIL:"
    ws["F63"] = "ITEM CAIXA:"
    for col in ["A","C","F"]:
        ws[f"{col}63"].font = _font(bold=True)

    # ── Aprovações ────────────────────────────────────────────
    ws["A65"] = "Aprovação de Medição"
    ws["A65"].font = _font(bold=True)

    ws["A66"] = "Data:"
    ws["C66"] = "Data:"
    ws["F66"] = "Data:"
    ws["A68"] = "CONTRATANTE"
    ws["C68"] = "CONTRATANTE"
    ws["F68"] = "CONTRATANTE"
    ws["A69"] = "Companhia Riograndense de Saneamento - CORSAN"
    ws["C69"] = "Coordenador/Gerente da Área"
    ws["F69"] = "Representante Legal da Unidade"

    for r in [68, 69]:
        for col in ["A","C","F"]:
            ws[f"{col}{r}"].font = _font(bold=(r==68), size=9)
            ws[f"{col}{r}"].alignment = _align(h="center")


def _criar_aba_boletim(wb: openpyxl.Workbook):
    ws = wb.create_sheet("BOLETIM")

    # Larguras
    for col, w in [("A",6),("B",35),("C",4),("D",6),("E",12),
                   ("F",10),("G",14),("H",14),("I",14),
                   ("J",14),("K",14),("L",14),("M",14)]:
        ws.column_dimensions[col].width = w

    cinza = _fill("D9D9D9")
    thin  = _thin_border()

    # ── Linha 1: título (mesclado A1:M1) ─────────────────────
    ws.merge_cells("A1:M1")
    ws["A1"].font      = _font(bold=True, size=13)
    ws["A1"].alignment = _align(h="center")
    ws["A1"].fill      = cinza

    # ── Linha 2: empresa / contrato / modalidade ──────────────
    ws["D2"] = "Empresa:"
    ws["H2"] = "Contrato :"
    ws["L2"] = "Modalidade :"
    for cell in ["D2","H2","L2"]:
        ws[cell].font      = _font(bold=True)
        ws[cell].alignment = _align()

    # ── Linha 3: local / período ──────────────────────────────
    ws["D3"] = "Local:"
    ws["H3"] = "Período :"
    for cell in ["D3","H3"]:
        ws[cell].font      = _font(bold=True)
        ws[cell].alignment = _align()

    # ── Linha 4: data base ────────────────────────────────────
    ws["D4"] = "Data Base :"
    ws["D4"].font      = _font(bold=True)
    ws["D4"].alignment = _align()

    # ── Linha 5: cabeçalho da tabela ─────────────────────────
    headers_row5 = {
        1: "ITEM", 2: "DESCRIÇÃO", 4: "UND",
        5: "QUANT.", 6: "P.U.", 7: "PREVISTO",
        8: "Q U A N T I D A D E S",
        11: "V A L O R E S"
    }
    for col, label in headers_row5.items():
        c = ws.cell(row=5, column=col, value=label)
        c.font      = _font(bold=True, size=9)
        c.alignment = _align(h="center")
        c.fill      = cinza
        c.border    = thin

    # ── Linha 6: subheader quantidades/valores ────────────────
    subheaders = {
        6: "$", 7: "$",
        8: "ACUM. ANT.", 9: "MES", 10: "ACUM. TOTAL",
        11: "ACUM. ANT.", 12: "MES", 13: "ACUM. TOTAL"
    }
    for col, label in subheaders.items():
        c = ws.cell(row=6, column=col, value=label)
        c.font      = _font(bold=True, size=9)
        c.alignment = _align(h="center")
        c.fill      = cinza
        c.border    = thin

    # ── Linha 7: linha do item (preenchida dinamicamente) ─────
    for col in range(1, 14):
        ws.cell(row=7, column=col).border = thin

    # ── Linhas 8–31: linhas vazias da tabela ─────────────────
    for row in range(8, 32):
        for col in [7, 10, 11, 12, 13]:
            c = ws.cell(row=row, column=col, value=0)
            c.border = thin

    # ── Linha 32: separador ───────────────────────────────────
    # (vazia)

    # ── Linha 33: TOTAL DO VALOR DO CONTRATO ─────────────────
    ws.cell(row=33, column=2,  value="TOTAL DO VALOR DO CONTRATO")
    ws.cell(row=33, column=2).font = _font(bold=True)
    ws.cell(row=33, column=2).fill = cinza

    # ── Linha 35: MEDIÇÃO BRUTA ───────────────────────────────
    ws.cell(row=35, column=2, value="MEDIÇÃO BRUTA (R$)")
    ws.cell(row=35, column=2).font = _font(bold=True)
    ws.cell(row=35, column=2).fill = cinza

    # ── Linha 36: FATURAMENTO DIRETO ─────────────────────────
    ws.cell(row=36, column=2, value="FATURAMENTO DIRETO (R$)")
    ws.cell(row=36, column=2).font = _font(bold=True)
    ws.cell(row=36, column=2).fill = cinza

    # ── Linha 37: ADIANTAMENTO ───────────────────────────────
    ws.cell(row=37, column=2, value="ADIANTAMENTO")
    ws.cell(row=37, column=2).font = _font(bold=True)
    ws.cell(row=37, column=2).fill = cinza

    # ── Linha 38: MEDIÇÃO LÍQUIDA ─────────────────────────────
    ws.cell(row=38, column=2, value="MEDIÇÃO LÍQUIDA (R$)")
    ws.cell(row=38, column=2).font = _font(bold=True)
    ws.cell(row=38, column=2).fill = cinza

    # Bordas e zeros nas linhas de totalização
    for row in [33, 35, 36, 37, 38]:
        for col in [7, 11, 12, 13]:
            c = ws.cell(row=row, column=col)
            c.border = thin
            if c.value is None:
                c.value = 0

    # ── Aprovações ────────────────────────────────────────────
    ws.cell(row=40, column=2,  value="Aprovação de Medição")
    ws.cell(row=40, column=2).font = _font(bold=True)

    for col, label in [(1,"Data:"),(3,"Data:"),(6,"Data:"),(8,"Data:"),(11,"Data:")]:
        ws.cell(row=41, column=col, value=label)

    for col, label in [
        (1,"4C DIGITAL"),(3,"CONTRATANTE"),(6,"CONTRATANTE"),
        (8,"CONTRATANTE"),(11,"CONTRATANTE")
    ]:
        ws.cell(row=43, column=col, value=label)
        ws.cell(row=43, column=col).font = _font(bold=True)

    for col, label in [
        (1,"Representante Legal"),
        (3,"Companhia Riograndense de Saneamento - CORSAN"),
        (6,"Supervisor/Coordenador Área"),
        (8,"Gerente da Área"),
        (11,"Representante Legal da Unidade")
    ]:
        ws.cell(row=44, column=col, value=label)
        ws.cell(row=44, column=col).font = _font(size=9)


def garantir_modelo():
    """
    Garante que o Modelo_medio.xlsx existe.
    Se não existir, cria do zero com a estrutura correta.
    """
    if MODELO_FILE.exists():
        return

    MODELO_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    _criar_aba_protocolo(wb)
    _criar_aba_boletim(wb)

    wb.save(MODELO_FILE)


# ─────────────────────────────────────────────────────────────
# PREENCHIMENTO DA MEDIÇÃO
# ─────────────────────────────────────────────────────────────

def gerar_excel_medicao(dados: dict) -> Path:
    """
    Garante o modelo, copia e preenche com os dados da medição.
    Retorna o Path do arquivo gerado.
    """

    # Garante que o modelo existe (cria se necessário)
    garantir_modelo()

    # ── Desempacota ───────────────────────────────────────────
    contrato          = str(dados["contrato"])
    num_medicao       = int(dados["num_medicao"])
    periodo           = dados["periodo"]
    data_apresentacao = dados["data_apresentacao"]
    descricao_servico = dados["descricao_servico"]
    valor_mes         = float(dados["valor_mes"])
    quant_mes         = float(dados["quant_mes"])

    empresa           = dados["empresa"]
    local             = dados["local"]
    modalidade        = dados["modalidade"]
    data_base         = dados["data_base"]
    data_termino      = dados["data_termino"]
    valor_original    = float(dados["valor_original"])
    item_num          = dados["item_num"]
    und               = dados["und"]
    quant_total       = float(dados["quant_total"])
    preco_unitario    = float(dados["preco_unitario"])
    centro_custo      = str(dados.get("centro_custo", ""))
    conta_contabil    = str(dados.get("conta_contabil", ""))
    item_caixa        = str(dados.get("item_caixa", ""))

    quant_acum_ant    = float(dados["quant_acum_ant"])
    valor_acum_ant    = float(dados["valor_acum_ant"])
    quant_acum_total  = float(dados["quant_acum_total"])
    valor_acum_total  = float(dados["valor_acum_total"])

    historico         = dados.get("historico", [])

    mes_execucao = extrair_mes_do_periodo(periodo)

    # ── Nome do arquivo de saída ──────────────────────────────
    mes_slug = mes_execucao.replace(" ", "_")
    nome = f"medicao_{num_medicao:02d}_{contrato}_{mes_slug}.xlsx"
    output_path = OUTPUT_DIR / nome

    shutil.copy(MODELO_FILE, output_path)
    wb = openpyxl.load_workbook(output_path)

    # ══════════════════════════════════════════════════════════
    # ABA PROTOCOLO
    # ══════════════════════════════════════════════════════════
    ws = wb["PROTOCOLO"]

    # Variáveis da medição
    _w(ws, "A6",  f"Boletim de Medição n. {num_medicao:02d}")
    _w(ws, "A7",  descricao_servico)
    _w(ws, "B12", periodo)
    _w(ws, "C14", mes_execucao)
    _w(ws, "C15", _fmt_date(data_apresentacao))

    # Fixos do contrato
    _w(ws, "D3",  contrato)
    _w(ws, "B8",  empresa)
    _w(ws, "B9",  contrato)
    _w(ws, "H9",  modalidade)
    _w(ws, "B10", local)
    _w(ws, "B11", _fmt_date(data_base))
    _w(ws, "H12", _fmt_date(data_termino))
    _w(ws, "H17", valor_original)
    _w(ws, "H18", valor_original)
    _w(ws, "H23", valor_original)
    _w(ws, "H25", valor_original)

    # ── Bloco amarelo: histórico de boletins ──────────────────
    LINHA_HIST_INICIO = 30

    # Limpa bloco anterior
    for r in range(LINHA_HIST_INICIO, LINHA_HIST_INICIO + 60):
        ws.cell(row=r, column=4, value=None)
        ws.cell(row=r, column=8, value=None)

    # Escreve histórico
    for i, h in enumerate(historico):
        row = LINHA_HIST_INICIO + i
        c_label = ws.cell(row=row, column=4, value=h["label"])
        c_label.font      = _font(size=10)
        c_label.alignment = _align()

        c_valor = ws.cell(row=row, column=8, value=h["valor"])
        c_valor.font         = _font(size=10)
        c_valor.alignment    = _align(h="right")
        c_valor.number_format = '#,##0.00'

    n_boletins        = len(historico)
    ultima_linha_hist = LINHA_HIST_INICIO + n_boletins - 1

    # H28: soma dinâmica do histórico
    if n_boletins > 0:
        _w(ws, "H28", f"=SUM(H{LINHA_HIST_INICIO}:H{ultima_linha_hist})")
    else:
        _w(ws, "H28", 0)

    # ── Linhas calculadas (dinâmicas após o histórico) ────────
    # A estrutura fixa abaixo do bloco é sempre:
    # ultima_linha_hist + 1  → linha vazia
    # ultima_linha_hist + 2  → 5.- DESCONTO DE MATERIAIS   (fixo no modelo: 46)
    # ultima_linha_hist + 3  → vazia
    # ultima_linha_hist + 4  → vazia
    # ultima_linha_hist + 5  → vazia
    # ultima_linha_hist + 6  → vazia
    # ultima_linha_hist + 7  → 6.- RETENÇÕES               (fixo no modelo: 51)
    # ultima_linha_hist + 8  → cabeçalho retenções
    # ultima_linha_hist + 9  → vazia
    # ultima_linha_hist + 10 → 7.- OBSERVAÇÃO
    # ultima_linha_hist + 11 → vazia
    # ultima_linha_hist + 12 → vazia
    # ultima_linha_hist + 13 → VALOR BRUTO DA MEDIÇÃO NO MÊS  ← H aqui
    # ultima_linha_hist + 14 → vazia
    # ultima_linha_hist + 15 → SALDO DO CONTRATO JURÍDICO     ← H aqui
    # ultima_linha_hist + 16 → vazia
    # ultima_linha_hist + 17 → SALDO DA SOURCING              ← H aqui
    # ultima_linha_hist + 18 → vazia
    # ultima_linha_hist + 19 → CENTRO DE CUSTO / CC / IC      ← A,C,F,H aqui

    # Com 15 boletins: ultima = 44 → 44+13=57 ✓ (bate com o modelo original)

    l_bruto    = ultima_linha_hist + 13
    l_saldo_j  = ultima_linha_hist + 15
    l_saldo_s  = ultima_linha_hist + 17
    l_cc       = ultima_linha_hist + 19

    _w(ws, f"H{l_bruto}",   valor_mes)
    _w(ws, f"H{l_saldo_j}", "=H17-H28")
    _w(ws, f"H{l_saldo_s}", "=H17-H28")

    ws[f"H{l_bruto}"].number_format   = '#,##0.00'
    ws[f"H{l_saldo_j}"].number_format = '#,##0.00'
    ws[f"H{l_saldo_s}"].number_format = '#,##0.00'

    # Labels do centro de custo (preserva os que já existem no modelo)
    _w(ws, f"A{l_cc}", f"CENTRO DE CUSTO: {centro_custo}")
    _w(ws, f"C{l_cc}", f"CONTA CONTÁBIL: {conta_contabil}")
    _w(ws, f"F{l_cc}", f"ITEM CAIXA: {item_caixa}")
    _w(ws, f"H{l_cc}", valor_mes)
    ws[f"H{l_cc}"].number_format = '#,##0.00'

    # ══════════════════════════════════════════════════════════
    # ABA BOLETIM
    # ══════════════════════════════════════════════════════════
    ws_b = wb["BOLETIM"]

    # Linha 1: título
    _w(ws_b, "A1", f"BOLETIM DE MEDIÇÃO Nº {num_medicao:02d} - {mes_execucao}")

    # Linha 2: empresa / contrato / modalidade
    _w(ws_b, "E2", empresa)
    _w(ws_b, "I2", contrato)
    _w(ws_b, "M2", modalidade)

    # Linha 3: local / período
    _w(ws_b, "E3", local)
    _w(ws_b, "I3", periodo)

    # Linha 4: data base
    _w(ws_b, "E4", _fmt_date(data_base))

    # Linha 7: dados do item
    # (no modelo o cabeçalho ocupa linhas 5–6, item começa na 7)
    LINHA_ITEM = 7
    _wn(ws_b, LINHA_ITEM, 1,  item_num)
    _wn(ws_b, LINHA_ITEM, 2,  descricao_servico)
    _wn(ws_b, LINHA_ITEM, 4,  und)
    _wn(ws_b, LINHA_ITEM, 5,  quant_total)
    _wn(ws_b, LINHA_ITEM, 6,  preco_unitario)
    _wn(ws_b, LINHA_ITEM, 7,  valor_original)
    _wn(ws_b, LINHA_ITEM, 8,  quant_acum_ant)
    _wn(ws_b, LINHA_ITEM, 9,  quant_mes)
    _wn(ws_b, LINHA_ITEM, 10, quant_acum_total)
    _wn(ws_b, LINHA_ITEM, 11, valor_acum_ant)
    _wn(ws_b, LINHA_ITEM, 12, valor_mes)
    _wn(ws_b, LINHA_ITEM, 13, valor_acum_total)

    # Formata valores como número
    for col in [5, 6, 7, 8, 9, 10, 11, 12, 13]:
        ws_b.cell(row=LINHA_ITEM, column=col).number_format = '#,##0.00'

    # Linhas de totalização
    # TOTAL DO VALOR DO CONTRATO → linha 33
    _wn(ws_b, 33, 7,  valor_original)
    _wn(ws_b, 33, 11, valor_acum_ant)
    _wn(ws_b, 33, 12, valor_mes)
    _wn(ws_b, 33, 13, valor_acum_total)

    # MEDIÇÃO BRUTA → linha 35
    _wn(ws_b, 35, 11, valor_acum_ant)
    _wn(ws_b, 35, 12, valor_mes)
    _wn(ws_b, 35, 13, valor_acum_total)

    # FATURAMENTO DIRETO → linha 36
    for col in [11, 12, 13]:
        _wn(ws_b, 36, col, 0)

    # ADIANTAMENTO → linha 37
    for col in [11, 12, 13]:
        _wn(ws_b, 37, col, 0)

    # MEDIÇÃO LÍQUIDA → linha 38
    _wn(ws_b, 38, 11, valor_acum_ant)
    _wn(ws_b, 38, 12, valor_mes)
    _wn(ws_b, 38, 13, valor_acum_total)

    # Formata totais
    for row in [33, 35, 36, 37, 38]:
        for col in [7, 11, 12, 13]:
            ws_b.cell(row=row, column=col).number_format = '#,##0.00'

    wb.save(output_path)
    return output_path
